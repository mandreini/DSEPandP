# -*- coding: utf-8 -*-
"""
Created on Wed Dec 12 14:28:58 2018

@author: mandreini
"""
import numpy as np
import matplotlib.pyplot as plt
import openpyxl

class Engine(object):
    
    def __init__(self, P_sl, name='not given', config_used=None, mass=None, sfc=None, size=None):
        '''
        Create an engine object
        To be changed as i.e. Max does the engine work, this is a basic structure
        
        Parameters
        ----------
        P_sl : float
            Power available at sea level
        weight : float
            Weight of the engine, in kg
        sfc : float
            Specific fuel consumption of the engine, in N/s
        '''
        
        self.P_sl = P_sl
        self.mass = mass
        self.sfc = sfc
        self.size = size
        self.name = name
        self.config_used = config_used
        
    @staticmethod
    def get_rho(h):
        return 1.225*288.16/(15+271.16-1.983*h/304.8)*(1-(0.001981*h/0.3048)/288.16)**5.256
    
    @staticmethod
    def get_T(h):
        return 298.13 + -0.0065*h
        
    def determine_power_at_altitude(self, h):
        '''
        Function to determine power provided by the engine at given altitude
        Power available decreases 1:1 with density decrease
        Power available increases 0.1% per degree K temperature decrease
        
        Parameters
        ----------
        altitude : array_like
            (An array of) operating altitude(s)
        
        Returns
        -------
        array_like :
            power provided at given altitude(s)
        '''
        rhoa = self.get_rho(h)
        rho0 = self.get_rho(0)
        Ta = self.get_T(h)
        T0 = self.get_T(0)
        
        return self.P_sl*(rhoa/rho0)*(1+(T0-Ta)*0.001)
    
class EngineSpider(object):
    '''
    Object containing relevant engine data. To be replaced with a scrappy method when I want to
    '''
    
    def __init__(self, fname):
        # bootlegged spider
        engine_opts = []
        
        engine_data = openpyxl.load_workbook('DSE Propulsion HAMRAC.xlsx').get_sheet_by_name('Engines')
        
        row = 2
        while engine_data.cell(row, 1).value is not None:
            manufacturer = engine_data.cell(row, 1).value
            engine_type = engine_data.cell(row, 2).value
            power = engine_data.cell(row, 3).value
            weight = engine_data.cell(row, 5).value
            length = engine_data.cell(row, 6).value
            diameter = engine_data.cell(row, 8).value
            specific_fuel_consumption = engine_data.cell(row, 11).value
            
            engine = Engine(power, name=manufacturer+engine_type, mass=weight, sfc=specific_fuel_consumption, size=(length, diameter))
            engine_opts.append(engine)
            
            row += 1  # rip non-infinite loops
            
        
        '''
        edata = np.genfromtxt(fname, dtype=str, delimiter=',', usecols=(0, 1, 6, 13, 22, 23), skip_header=1)
        for dataline in edata:
            # if data is not given, use None
            try:
                psl = float(dataline[2])
            except:
                failed_ac.append(dataline)
                continue
            
            try:
                m = float(dataline[4])
            except:
                m = None
            
            try:
                fuelflow = float(dataline[3])
            except:
                fuelflow = None
                
            n = dataline[0].strip('"')
                
            engine = Engine(P_sl=psl, name=n, config_used=dataline[1], mass=m, sfc=fuelflow, size=None)
            engine_opts.append(engine) 
        
        self.failed = failed_ac 
        '''
        self.engines = engine_opts

    def add_engine(self, engine):
        ''' Manually add an engine to the engines property '''
        self.engines.append(engine)
        
    def select_engine(self, P_hov, h):
        '''
        Select the engine that meets the given power requirements providing the smallest power
        
        Parameters
        ----------
        P_hov : float
            The hover power, in W
        h : float
            The altitude of the engine for regulation 1 (TO)
            
        Returns
        -------
        best : Engine object    
            The engine with the available power closest to the hover requirements 
            that meets them
        possibilities : list; Engine objects
            All engines that meet the hover power requirements
        '''
        P_hov /= 1000
        possibilities = [i for i in self.engines if i.determine_power_at_altitude(h) > P_hov]
        best = Engine(99999)
        for p in possibilities:
            # smallest difference, (assume) all possibilities meet hover requirements
            if p.P_sl - P_hov < best.P_sl - P_hov:
                best = p
            
        return best, possibilities
    
    def estimate_mass(self, engine):
        '''
        Estimate the expected mass of an engine (that does not have a mass value given to it)
        based on the trend of the given masses of the other engines in the Spider
        Report: Data given on the drive is bad
        Proposal: Fix the data 
        
        Parameters
        ----------
        engine : Engine object
            Defined engine object
        '''
        if engine.mass:
            return engine.mass
        
        masses = []
        powers = []
        for e in self.engines:
            if e.mass:
                masses.append(e.mass)
                powers.append(e.P_sl)

        pminterp = np.interp(engine.P_sl, masses, powers)
        
        return pminterp, masses, powers
    
lbshrhp2nsw = lambda x: 0.453592*0.277778*9.81/0.745699872*1000 * x   # lbs/hr/hp to N/s/W... I think 
   
HoneywellT35 = Engine(1175, name='AH-1J SeaCobra', mass=312, sfc=lbshrhp2nsw(0.65), size=(120/9, 24.5))  # https://aerospace.honeywell.com/en/products/engines/t53-turboshaft-engine, 
# https://aerospace.honeywell.com/en/~/media/aerospace/files/brochures/n61-1733-000-000-t53-v6-bro.pdf, https://en.wikipedia.org/wiki/Lycoming_T53
KlimovVK2500 = Engine(1985, name='Kamov Ka-50', mass=300, sfc=0.22*3.6, size=(205.5, 72.8))  # https://en.wikipedia.org/wiki/Klimov_VK-2500

#print(HoneywellT35.determine_power_at_altitude(9000))

#engineoptions = EngineSpider('Aircraft Statistics - Sheet1.csv')
engineoptions = EngineSpider('DSE Propulsion HAMRAC.xlsx')

# manually add engines not found in the datasheet, some seem different to the ones in the sheet, so I left that alone
# also this often just considers 1 engine
B206BR = Engine(313, 'Bell 206BR', 'Çonventional', size=(0.99, 48.26), mass=72, sfc=1)
B212 = Engine(1980*0.74569987, 'Bell 212', 'Conventional', size=(66*2.54, 32.5*2.54))
B230 = B206BR
B407 = Engine(610, 'Bell 407GXP', 'Conventional', size=(0.92, 0.65), mass=153)
ATwinstar = B206BR
LeonardoAW119KX = B407
BChinook = Engine(3631, 'Boeing Chinook CH47-F (Block 2)', 'Tandem', size=(1.196, 0.615), mass=377)

#[engineoptions.add_engine(i) for i in [B206BR, B212, B230, B407, ATwinstar, LeonardoAW119KX, BChinook]]
